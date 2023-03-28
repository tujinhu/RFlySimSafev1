import threading
import numpy as np
import re
import sqlite3
import os
import sys
import json
from fnmatch import fnmatch
import matplotlib.pyplot as plt
import shutil
import pandas as pd
from openpyxl import Workbook
import AutoREG


def dict_factory(cursor, row):  
    d = {}  
    for idx, col in enumerate(cursor.description):  
        d[col[0]] = row[idx]  
    return d 

def ClearInd():
    DataAPImanger.mkFolderNum_quad = 0
    DataAPImanger.mkFolderNum_fixed = 0
    DataAPImanger.mkFolderNum_usv = 0
    DataAPImanger.Ind = 0 
    DataAPImanger.Ind_quad = 0
    DataAPImanger.Ind_fixed = 0
    DataAPImanger.Ind_usv = 0
    DataAPImanger.deleteflag_quad = 0
    DataAPImanger.deleteflag_fixed = 0
    DataAPImanger.deleteflag_usv = 0
    DataAPImanger.TrueDataOvercount = []

class mavdb:
    def __init__(self,TestBatPath):
        self.cursor = 0
        self.mydb = 0
        self.TestBatPath = TestBatPath
        self.is_tested = 0
        self.count = 0
        self.jsonpath = os.path.join(sys.path[0],'..','Model',TestBatPath[1],'db.json')
        mavdb.JSON_TO_SQL(self)
        self.VISIONFLAG = mavdb.Vision(self)

    def Vision(self):
        jsonpath = self.jsonpath
        with open(jsonpath, "r",encoding='utf-8') as f:
            db_data = json.load(f)
        isVision = db_data.get('Vision') 
        if isVision == 'on':
            return True
        else:
            return False

    def GET_CURSOR(self):
        imPath = os.path.join(sys.path[0],'..','fault.db')
        self.mydb = sqlite3.connect(imPath)
        self.mydb.row_factory = dict_factory
        self.cursor=self.mydb.cursor()

    def GET_FAULT_CASE(self): # Obtain fault test cases
        # 获取游标
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: # Quad database
            sql='''
            select   *
            from     Quadrotorfaultcase
            '''
        self.cursor.execute(sql)
        result=self.cursor.fetchall()
        return result

    def GET_CASEID(self): # Obtain the list of fault test case IDs
        result = mavdb.GET_FAULT_CASE(self)
        path = self.jsonpath

        with open(path, "r",encoding='utf-8') as f:
            db_data = json.load(f)

        global caselist
        if db_data.get('testcase') == 'all':
            caselist = []
            casedata = result
            for data in casedata:
                ID = data.get('CaseID')
                caselist.append(ID)
            return caselist
        else:
            if AutoREG.REG.FRAME_TYPE == 1: # Single Type
                if len(re.findall(r';',db_data.get('testcase')))>=1 and AutoREG.REG.MAV_ALL_NUM >=2: # Single machine multiple instances
                    MultiTemplist = []
                    caselist = [val for val in re.split(';',db_data.get('testcase'))]
                    for i in range(len(caselist)):
                        if len(caselist[i]) > 0:
                            Templist = [int(val) for val in re.split(',',caselist[i])]
                            MultiTemplist.append(Templist)
                    MultiMavCaselist = []
                    for i in range(len(MultiTemplist[0])):
                        Templist = []
                        for j in range(len(MultiTemplist)):
                            if len(MultiTemplist[j]) > 0:
                                Templist.append(MultiTemplist[j][i])
                        MultiMavCaselist.append(Templist)
                    return MultiMavCaselist
                elif AutoREG.REG.MAV_ALL_NUM >=2:
                    MultiMavCaselist = []
                    caselist = [int(val) for val in re.split(',',db_data.get('testcase'))] # eg: [3, 1, 30]
                    for i in range(len(caselist)):
                        clist = []
                        for j in range(self.mavNum):
                            clist.append(caselist[j])
                        MultiMavCaselist.append(clist)
                    return MultiMavCaselist
                else:
                    caselist = [[int(val) for val in re.split(',',db_data.get('testcase'))]] # [3, 1, 30]
                    return caselist
            else:
                if len(re.findall(r';',db_data.get('testcase')))>=1: # Multiple models and instances
                    MultiTemplist = []
                    caselist = [val for val in re.split(';',db_data.get('testcase'))]
                    for i in range(len(caselist)):
                        if len(caselist[i]) > 0:
                            Templist = [int(val) for val in re.split(',',caselist[i])]
                            MultiTemplist.append(Templist)
                    MultiMavCaselist = []
                    for i in range(len(MultiTemplist[0])):
                        Templist = []
                        for j in range(len(MultiTemplist)):
                            if len(MultiTemplist[j]) > 0:
                                Templist.append(MultiTemplist[j][i])
                        MultiMavCaselist.append(Templist)
                    return MultiMavCaselist
                else: # Single instance of multiple models
                    caselist = [int(val) for val in re.split(',',db_data.get('testcase'))] # [3, 1, 30]
                    return caselist

    def GET_MAVCMD(self,case_id): # Process Control Sequence
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            select * from Quadrotorfaultcase
            where CaseID = {}
            '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        case_sequence = data[0].get('ControlSequence')
        case = re.split(';',case_sequence)
        cmd = np.array([])
        for i in range(len(case)):
            cmd = np.append(cmd,case[i])
        return cmd

    def RESULT_DBPro(self,data): # Process test result library (add test results)
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            insert into Quadrotortestresult
            (CaseID,FaultTestType,FaultParameters,TestStatus,SafetyAssessmentResults)
            values(?,?,?,?,?) 
            '''
        value = (data[0],data[1],data[2],data[3],data[4])
        self.cursor.execute(sql,value)
        self.mydb.commit()
    
    def TEST_STATEPro(self,case_id): # Process test case library (change test status)
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            update Quadrotorfaultcase
            set TestStatus = 'Finished'
            where CaseID = {}
            '''.format(case_id)
        self.cursor.execute(sql)
        self.mydb.commit()

    def IS_TheSame_IDPro(self,case_id): # Determine whether a duplicate CaseID is inserted
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            select * from Quadrotorfaultcase
            where CaseID = {}
            '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
    
        if len(data) != 0:
            if self.TestBatPath[0] == 1: 
                sql2 = '''
                delete from Quadrotorfaultcase
                where CaseID = {}
                '''.format(case_id)
            self.cursor.execute(sql2)
            self.mydb.commit()
 
    def IS_TESTEDPro(self,case_id): # Judge whether it is a tested case
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            select * from Quadrotorfaultcase
            where CaseID = {}
            '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()

        if data[0].get('TestStatus') == 'Finished':
            self.is_tested = 1
        else:
            self.is_tested = 0
        return self.is_tested

    def RESETR_DB(self,case_id): # Handle the result library of repeated use case test
        mavdb.GET_CURSOR(self)
        if self.TestBatPath[0] == 1: 
            sql = '''
            select * from Quadrotorfaultcase
            where CaseID = {}
            '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()

        if len(data) != 0:
            if self.TestBatPath[0] == 1: 
                sql2 = '''
                delete from Quadrotortestresult
                where CaseID = {}
                '''.format(case_id)
            self.cursor.execute(sql2)
            self.mydb.commit()

    def JSON_TO_SQL(self):
        path = self.jsonpath
        db_case = mavdb.GET_FAULT_CASE(self)

        with open(path, "r",encoding='utf-8') as f:
            db_data = json.load(f)

        json_case = db_data.get('faultcase')

        # Find and compare use cases with the same CaseID, cover different use cases with the same ID, and add the remaining use cases in json to the use case library
        # 1、Extract the CaseID of json
        if len(json_case) >= 1:
            self.json_ID = []
            for i in range(len(json_case)):
                self.json_ID.append(json_case[i].get("CaseID"))

            # 2、Extract CaseID from database
            self.db_ID = []
            for i in range(len(db_case)):
                self.db_ID.append(db_case[i].get("CaseID"))

            self.con_ID = list(set(self.json_ID)&set(self.db_ID))
            con_json_case = []
            for i in range(len(json_case)):
                if json_case[i].get('CaseID') in self.con_ID:
                    con_json_case.append(json_case[i])

            con_db_case = []
            for i in range(len(db_case)):
                if db_case[i].get('CaseID') in self.con_ID:
                    con_db_case.append(db_case[i])

            # Compare whether the cases of the two are the same. If they are the same, the use cases in the use case library will be overwritten
            for i in range(len(con_json_case)):
                # If the data of json is not equal to the data in the database, the data in the database will be overwritten
                if con_db_case[i] != con_json_case[i]:
                    Testcase = []
                    for d in  con_json_case[i].items():
                        Testcase.append(d[1])
                    mavdb.GET_CURSOR(self)
                    Testcase = [str(var) for var in Testcase]
                    if self.TestBatPath[0] == 1: 
                        sql = '''
                            insert into Quadrotorfaultcase
                            (CaseID,Subsystem,FaultType,ControlSequence,InterestedLog,TestStatus)
                            values(?,?,?,?,?,?)
                        '''
                    values = Testcase
                    # If there is a CaseID primary key conflict in the use case library, delete the in the use case library and add the
                    mavdb.IS_TheSame_IDPro(self,values[0])
                    self.cursor.execute(sql,values)
                    self.mydb.commit()
                
            # Add the remaining cases in json_case into the use case library
            discon_jsonID = list(set(self.json_ID) - set(self.con_ID))
            discon_json_case = []
            for i in range(len(json_case)):
                if json_case[i].get('CaseID') in discon_jsonID:
                    discon_json_case.append(json_case[i])
            for i in range(len(discon_json_case)):
                Testcase = []
                for d in  discon_json_case[i].items():
                    Testcase.append(d[1])
                mavdb.GET_CURSOR(self)
                Testcase = [str(var) for var in Testcase]
                if self.TestBatPath[0] == 1: 
                    sql = '''
                        insert into Quadrotorfaultcase
                        (CaseID,Subsystem,FaultType,ControlSequence,InterestedLog,TestStatus)
                        values(?,?,?,?,?,?)
                    '''
                values = Testcase

                mavdb.IS_TheSame_IDPro(self,values[0])
                self.cursor.execute(sql,values)
                self.mydb.commit()

    def DataAPImanger(func):
        def wrapFunc(self,*args, **kwargs):
            self.count += 1
            return func(self,*args, **kwargs)
        return wrapFunc

    @DataAPImanger  
    def MAV_JSONPro(self,case_id): # Change the test status information of json file
        if self.count == 1: 
            path = self.jsonpath
            with open(path, "r",encoding='utf-8') as f:
                db_data = json.load(f)
            json_case = db_data.get('faultcase')

            if len(json_case) >= 1:
                if case_id in self.json_ID:
                    path = self.jsonpath

                    with open(path, "r",encoding='utf-8') as f:
                        db_data = json.load(f)
                        db_data['faultcase'][case_id-1]['TestStatus'] = 'Finished'
                        data = db_data
                    f.close()

                    with open(path, "w",encoding='utf-8') as w:
                        json.dump(data,w,indent=4) 
                    w.close()

class DataAPImanger:
    mkFolderNum_quad = []
    mkFolderNum_fixed = []
    mkFolderNum_usv = []
    Ind = 0
    Ind_quad = 0
    Ind_fixed = 0
    Ind_usv = 0
    deleteflag_quad = 0
    deleteflag_fixed = 0
    deleteflag_usv = 0
    lock = threading.Lock() 
    TrueDataOvercount = []
    
class DataAPI(DataAPImanger): 
    def __init__(self,CaseID,TestBatPath,mavInfo,Data):
        self.jsonpath = os.path.join(sys.path[0],'..','Model',TestBatPath[1],'db.json')
        self.Data = Data
        self.TestNumCount = mavInfo[1]
        self.TestBatPath = TestBatPath
        self.isTrueDataRecordOver = 0
        self.MacVechileNum = self.GetMavInsNum()

        pathStr = os.path.join(sys.path[0],'..','Model',TestBatPath[1],TestBatPath[2])
        
        self.PX4Path='C:\\PX4PSP'
        with open(pathStr, 'r',encoding='UTF-8') as file:
            for line in file:
                if line.find('SET PSP_PATH=')!=-1:
                    self.PX4Path=line.replace('SET PSP_PATH=','')
                    self.PX4Path=self.PX4Path.strip()
                    self.PX4Path=self.PX4Path.replace('\\','/')
                    break
        self.PlatFormpath = self.PX4Path 
        self.ProjectPath = os.path.dirname(os.path.dirname(__file__))
        self.DataPath = self.ProjectPath + '/Data'
        self.QuadDataPath = self.DataPath + '/Quadrotor'
        self.CaseID = CaseID
        self.lock.acquire() 
        self.MakeFolder()
        self.lock.release() 

    def GetMavInsNum(self):
        pathStr = os.path.join(sys.path[0],'..','Model',self.TestBatPath[1],self.TestBatPath[2])
        MavVechileNum = 1
        with open(pathStr, 'r',encoding='UTF-8') as file:
            for line in file:
                if line.find('SET /A VehicleNum=')!=-1:
                    MavVechileNum=line.replace('SET /A VehicleNum=','')
                    MavVechileNum=MavVechileNum.strip()
                    MavVechileNum=int(MavVechileNum)
                    break
        if self.TestBatPath[0] == 1:
            DataAPImanger.mkFolderNum_quad = [val+1 for val in range(MavVechileNum)]
        return MavVechileNum

    def MakeFolder(self):
            # 1、Get destination folder path
            # Single instance of single machine or single instance of multiple machines:
            if AutoREG.REG.TEST_MODE == 1 or AutoREG.REG.TEST_MODE == 3: 
                if self.TestBatPath[0] == 1: # Quad
                    self.TargetFilder = self.QuadDataPath + '/TestCase_{}'.format(self.CaseID) 
                    self.TargetFilder_log = self.TargetFilder + '/log' 
                    self.TargetFilder_userdata = self.TargetFilder + '/mydata' # User data
                    self.TargetFilder_truedata = self.TargetFilder + '/TrueData' 
                    if AutoREG.REG.TEST_MODE == 3:
                        DataAPImanger.Ind += 1
                    else:
                        DataAPImanger.Ind = 1
            # Single machine multi-instance or multi-machine multi-instance
            elif AutoREG.REG.TEST_MODE == 2 or AutoREG.REG.TEST_MODE == 4: 
                if self.TestBatPath[0] == 1: # Quad
                    self.TargetFilder = self.QuadDataPath + '/TestCase_{}'.format(self.TestNumCount) 
                    self.mavPath = self.TargetFilder + '/mav{}'.format(DataAPImanger.mkFolderNum_quad[DataAPImanger.Ind_quad])
                    self.TargetFilder_log = self.mavPath + '/log' 
                    self.TargetFilder_userdata = self.mavPath + '/mydata' 
                    self.TargetFilder_truedata = self.mavPath + '/TrueData' 
                    DataAPImanger.Ind_quad += 1
                    DataAPImanger.Ind += 1

            # 2、delete the specified folder,only delete once
            if AutoREG.REG.TEST_MODE == 2 or AutoREG.REG.TEST_MODE == 4:
                if self.TestBatPath[0] == 1:
                    if os.path.exists(self.TargetFilder) and DataAPImanger.deleteflag_quad == 0: # If the destination folder exists, delete the reconstruction
                        shutil.rmtree(self.TargetFilder)
                        DataAPImanger.deleteflag_quad = 1
                    else:
                        DataAPImanger.deleteflag_quad = 1
                elif self.TestBatPath[0] == 2:
                    if os.path.exists(self.TargetFilder) and DataAPImanger.deleteflag_fixed == 0: # If the destination folder exists, delete the reconstruction
                        shutil.rmtree(self.TargetFilder)
                        DataAPImanger.deleteflag_fixed = 1
                    else:
                        DataAPImanger.deleteflag_fixed = 1
                elif self.TestBatPath[0] == 3:
                    if os.path.exists(self.TargetFilder) and DataAPImanger.deleteflag_usv == 0: # If the destination folder exists, delete the reconstruction
                        shutil.rmtree(self.TargetFilder)
                        DataAPImanger.deleteflag_usv = 1
                    else:
                        DataAPImanger.deleteflag_usv = 1
                os.makedirs(self.TargetFilder_log) 
                os.makedirs(self.TargetFilder_userdata) 
                os.makedirs(self.TargetFilder_truedata) 
            elif AutoREG.REG.TEST_MODE == 1 or AutoREG.REG.TEST_MODE == 3:
                if os.path.exists(self.TargetFilder): # If the destination folder exists, delete the reconstruction
                        shutil.rmtree(self.TargetFilder)

                os.makedirs(self.TargetFilder_log) 
                os.makedirs(self.TargetFilder_userdata) 
                os.makedirs(self.TargetFilder_truedata) 

            self.DataPro()
    
    def DataPro(self):
            # 1、List the directories under the file
            logInd = DataAPImanger.Ind
            log_path = self.PlatFormpath + '/Firmware/build/px4_sitl_default/instance_{}/log'.format(logInd)
            PlatForm_log_dirs = os.listdir(log_path) 
            log_data = PlatForm_log_dirs[len(PlatForm_log_dirs)-1]
            self.path = os.path.join(log_path,log_data) 
            dirs = os.listdir(self.path) 

            # 2、Get the latest ulg file
            ulg = dirs[len(dirs)-1]
            ulgPath = os.path.join(self.path,ulg) 
            
            # 3、Copy the ulg file to the log folder
            TargetPath_log = self.TargetFilder_log + '/{}'.format(ulg)
            shutil.copyfile(ulgPath, TargetPath_log) 

            # 4、Convert ulg to csv
            os.chdir(self.TargetFilder_log)
            cmd = "SET PATH="+self.PlatFormpath+"\\Python38;"+self.PlatFormpath+"\\Python38\\Scripts; && "
            cmd = cmd + "for %i in (*); do ulog2csv %i"
            os.system(cmd) 
            
            #打开数据文件夹窗口
            # os.system('explorer .')

            # 5、Keep the ulg file and delete the unwanted csv file
            #1) Load json and get the csv file of interest
            with open(self.jsonpath, "r",encoding='utf-8') as f:
                db_data = json.load(f)

            csvfiles = db_data.get('testdata') 
            csvfiles_keys = []
            for key in csvfiles.keys():
                csvfiles_keys.append(key)

            #2) Document processing
            csvdirs = os.listdir(self.TargetFilder_log)
            csvdirs.remove(csvdirs[0]) 
            for file in csvdirs:
                flag = 0
                for k in range(len(csvfiles_keys)):
                    if(fnmatch(file,csvfiles_keys[k])):
                        flag = 1
                if flag == 0:
                    fpath = os.path.join(self.TargetFilder_log,file) 
                    os.remove(fpath) 

            remaindirs = os.listdir(self.TargetFilder_log)
            for file in remaindirs:
                if(fnmatch(file,'*yaw_estimator_status_0.csv')):
                    os.remove(file) 
            
            self.GetData()

    def GetData(self): # Extract the logs and corresponding columns of interest to the user and merge them into the specified csv file
            # 1、Create a csv file in the specified path (user selected folder path)
            usercsvpath = self.TargetFilder_userdata + '/mycsv.csv'
            frame = pd.DataFrame() 
            frame.to_csv(usercsvpath,index=False)    

            # 2、Get the data of the log file selected by the user
            #1) Get the fault case of the database
            with open(self.jsonpath, "r",encoding='utf-8') as f:
                db_data = json.load(f)
            faultcase = db_data.get("faultcase")

            #2) Get the test case corresponding to the current CaseID, the log and column of interest
            for d in faultcase:
                if d.get("CaseID") == self.CaseID:
                    curcase = d
            InteststedLog = curcase.get("InterestedLog")
            interestcsv = [] # eg:['*actuator_outputs_0.csv', '*airspeed_0.csv']
            interestdata = [] # eg:[['timestamp', 'noutputs', 'output[0]', 'output[1]', 'output[2]', 'output[3]', 'output[4]', 'output[5]'], ['timestamp', 'indicated_airspeed_m_s', 'true_airspeed_m_s', 'air_temperature_celsius']]
            for csv in InteststedLog.items():
                interestcsv.append(csv[0]) # Keys: tables of interest
                interestdata.append(csv[1]) # Values: data of interest

            #3) Take out the table of interest and the corresponding data column
            #3.1) Get the csv file in the target log folder
            csvfiles = os.listdir(self.TargetFilder_log)
            targetcsv = interestcsv
            flag = 0
            for k in range(len(csvfiles)):
                for j in range(len(targetcsv)):
                    if(fnmatch(csvfiles[k],targetcsv[j])): # Match target csv file
                        index = j 
                        targetcsvpath = self.TargetFilder_log + '/{}'.format(csvfiles[k]) 
            #3.2) Open the interested table and extract the corresponding interested data to the user csv
                        usercols = interestdata[index] 
                        disereddatas = pd.read_csv(targetcsvpath,usecols=usercols) 
                        df2 = pd.DataFrame(disereddatas) 
                        if flag == 1:
                            df1 = pd.read_csv(usercsvpath) 
                            framedata = pd.concat([df1,df2],axis=1)
                            framedata["Table{}".format(j)] = targetcsv[j] 
                            framedata["{}".format(j)] = " " 
                        else:
                            framedata = df2
                            framedata["Table{}".format(j)] = targetcsv[j] 
                            framedata["{}".format(j)] = " " 

                        framedata.to_csv(usercsvpath,index=False) 
                        flag = 1

            self.TruedataRecord()

    def TruedataRecord(self):
            wb = Workbook()
            sheet1 = wb.create_sheet('v')
            sheet2 = wb.create_sheet('ang')
            sheet3 = wb.create_sheet('pos')
            mean_arr_v = self.Data[0]
            mean_arr_ang = self.Data[1]
            mean_arr_pos = self.Data[2]

            for i in range(len(mean_arr_v)):
                v = list(mean_arr_v[i])
                sheet1.append(v)
                ang = list(mean_arr_ang[i])
                sheet2.append(ang)
                pos = list(mean_arr_pos[i])
                sheet3.append(pos)
        
            truedata_xlsxpath = self.TargetFilder_truedata + '//truedata.xlsx'
            wb.save(truedata_xlsxpath)

            data_xls_v = pd.read_excel(truedata_xlsxpath,sheet_name='v')
            truedata_csvpath_v = self.TargetFilder_truedata + '//truedata_vel.csv' 
            data_xls_v.to_csv(truedata_csvpath_v, encoding='utf-8',index=False,header=None) 

            data_xls_ang = pd.read_excel(truedata_xlsxpath,sheet_name='ang')
            truedata_csvpath_ang = self.TargetFilder_truedata + '//truedata_ang.csv' 
            data_xls_ang.to_csv(truedata_csvpath_ang, encoding='utf-8',index=False,header=None) 

            data_xls_pos = pd.read_excel(truedata_xlsxpath,sheet_name='pos')
            truedata_csvpath_pos = self.TargetFilder_truedata + '//truedata_pos.csv' 
            data_xls_pos.to_csv(truedata_csvpath_pos, encoding='utf-8',index=False,header=None) 
            # os.remove(truedata_xlsxpath) 
            self.isTrueDataRecordOver = 1
            DataAPImanger.TrueDataOvercount.append(self.isTrueDataRecordOver)
    
    def ProfustSaftyScoreRecord(self,AssCount,IndexName,ProfustData):
        wb = Workbook()
        for i in range(len(IndexName)):
            sheet = wb.create_sheet('{}'.format(IndexName[i]))
            for j in range(len(ProfustData[i])):
                SaftyScore = [ProfustData[i][j]]
                sheet.append(SaftyScore)

        ProfustDataStagePath = self.TargetFilder_ProfustData + '/Stage{}'.format(AssCount)
        os.makedirs(ProfustDataStagePath) 
        RataModelSaftyData_xlsx_path = ProfustDataStagePath + '//ProfustData.xlsx'
        wb.save(RataModelSaftyData_xlsx_path)

        for i in range(len(IndexName)):
            RataModelSaftyData = pd.read_excel(RataModelSaftyData_xlsx_path,sheet_name='{}'.format(IndexName[i]))
            RataModelSaftyData_csv_path = ProfustDataStagePath + '//ProfustSaftyScore_{}.csv'.format(IndexName[i])
            RataModelSaftyData.to_csv(RataModelSaftyData_csv_path, encoding='utf-8',index=False)
        os.remove(RataModelSaftyData_xlsx_path)
    
    def DrawTrue(self,data):
        TrueDataPath_Pos = self.TargetFilder_truedata + '//truedata_{}.csv'.format(data)
        Data = pd.read_csv(TrueDataPath_Pos)
        Data = np.array(Data.values.tolist())
        fig = plt.figure()
        mydata = fig.add_subplot(111, projection='3d')
        mydata_x = Data[:,0]
        mydata_y = Data[:,1]
        mydata_z = abs(Data[:,2])
        mydata.plot(mydata_x,mydata_y,mydata_z,label='{}'.format(data))
        mydata.legend()
        fig.show()
        plt.pause(3)
        plt.close('all') 
    
    